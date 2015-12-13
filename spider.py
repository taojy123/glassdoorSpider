
import cookielib
import urllib2, urllib
import time
import re
import traceback
import time
import json
import gc
from openpyxl import Workbook
from openpyxl.reader.excel  import  load_workbook
from BeautifulSoup import BeautifulSoup

print "==============1109=================="


open('finish_cmp.txt', 'a')
open('finish_page.txt', 'a')
open('error_page.txt', 'a')

finish_cmp = open('finish_cmp.txt').read().strip()
finish_cmps = finish_cmp.split('\n')

finish_page = open('finish_page.txt').read().strip()
finish_pages = finish_page.split('\n')


cj = cookielib.CookieJar()
opener = urllib2.build_opener(urllib2.HTTPCookieProcessor(cj))
#opener = urllib2.build_opener(urllib2.HTTPCookieProcessor(cj), urllib2.ProxyHandler({'http':"10.239.120.37:911"}))
opener.addheaders = [
                    ('User-agent', 'Mozilla/5.0 (Windows NT 5.2) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/21.0.1180.89 Safari/537.1'),
                     ]
 
def get_page(url, data=None):
    print url
    resp = None
    n = 0
    while n < 10:
        n += 1
        try:
            resp = opener.open(url, data, timeout=10)
            page = resp.read()
            page = page.decode("gbk", "replace").encode("gbk", "replace")
            return page
        except:
            traceback.print_exc()
            time.sleep(2)
            print "Try after 2 seconds ..."
            continue
    raise Exception("Get page failed")


wb1 = load_workbook('input.xlsx')
wb2 = load_workbook('output.xlsx')
sheet1 = wb1.get_sheet_by_name(wb1.sheetnames[0])
sheet2 = wb2.get_sheet_by_name(wb2.sheetnames[0])


j = 2
while True:
    j += 1
    if not sheet2["A%d"%j].value:
        break

print 'start from:', j


# GSM_GD_RA1@126.com  /  pkugsm
url = "https://www.glassdoor.com/profile/ajax/loginAjax.htm?username=GSM_GD_RA1%40126.com&password=pkugsm"
p = get_page(url)
print p


try:
    i = 0
    while True:
        print "========================================"
        i += 1
        url = sheet1["C%d"%i].value
        if not url:
            break
        if "http" not in url:
            continue
        name = sheet1["B%d"%i].value
        stock_code = sheet1["A%d"%i].value

        if url in finish_cmps:
            print url, "exist!!!!!!!"
            continue

        # url = "http://www.glassdoor.com/Reviews/ZTE-Reviews-E40056.htm"
        # url = "http://www.glassdoor.com/Reviews/XCMG-Reviews-E662352.htm"
        # url = "http://www.glassdoor.com/Reviews/China-Life-Insurance-Reviews-E33262.htm"

        p = get_page(url)

        rs = re.findall(r"-E(\w+).htm", url)
        if not rs:
            print url, 'not found code!!!'
            continue
        code = rs[0]
        print code

        p_soup = BeautifulSoup(p)

        # name = p_soup.find("p", "h1 strong tightAll").getText()

        reviews = "N/A"
        a_soup = p_soup.find("a", "eiCell cell reviews active")
        if a_soup:
            span_soup = a_soup.find("span", "num h2 notranslate")
            reviews = span_soup.getText()

        try:
            print name, stock_code, reviews
        except:
            pass


        ajax_url = "http://www.glassdoor.com/api/employer/%s-rating.htm" % code
        p = get_page(ajax_url)
        overallRating = cultureAndValues = workLife = seniorManagement = compAndBenefits = careerOpportunities = "N/A"
        rating = json.loads(p)
        for r in rating.get("ratings", []):
            if r["type"] == "overallRating":
                overallRating = r["value"]
                overallRating = "%.1f" % overallRating
            if r["type"] == "cultureAndValues":
                cultureAndValues = r["value"]
                cultureAndValues = "%.1f" % cultureAndValues
            if r["type"] == "workLife":
                workLife = r["value"]
                workLife = "%.1f" % workLife
            if r["type"] == "seniorManagement":
                seniorManagement = r["value"]
                seniorManagement = "%.1f" % seniorManagement
            if r["type"] == "compAndBenefits":
                compAndBenefits = r["value"]
                compAndBenefits = "%.1f" % compAndBenefits
            if r["type"] == "careerOpportunities":
                careerOpportunities = r["value"]
                careerOpportunities = "%.1f" % careerOpportunities

        print overallRating, cultureAndValues, workLife, seniorManagement, compAndBenefits, careerOpportunities

        flag = False
        for k in range(1, 999999):
            try:
                gc.collect()

                p_url = url.replace(".htm", "_P%d.htm"%k)
                p_url += "?filter.defaultEmploymentStatuses=false&filter.employmentStatus=REGULAR&filter.employmentStatus=PART_TIME&filter.employmentStatus=CONTRACT&filter.employmentStatus=INTERN&filter.employmentStatus=FREELANCE&filter.employmentStatus=UNKNOWN"

                if p_url in finish_pages:
                    flag = True
                    print p_url, "========== exist =========="
                    continue

                p = get_page(p_url)
                p_soup = BeautifulSoup(p)

                comments = p_soup.findAll("div", "hreview")

                m = 0
                for comment in comments:

                    flag = True
                    m += 1
                    n = (k-1) * 10 + m

                    print n

                    time_soup = comment.find("time")
                    if time_soup:
                        comment_time = time_soup.get("datetime")
                    else:
                        comment_time = "N/A"
                    print "comment_time:", comment_time

                    span_soup = comment.find("span", "authorInfo tbl hideHH")
                    s = span_soup.getText()
                    index = s.find("-")
                    a = s[:index]
                    b = s[index+1:]
                    former_current = "N/A"
                    if "Current" in a:
                        former_current = "Current"
                    elif "Former" in a:
                        former_current = "Former"
                    job_title = b.strip()
                    print "former_current:", former_current

                    div_soup = comment.find("div", "cell reviewBodyCell")
                    working = div_soup.find("p", "notranslate").getText()
                    intern_full_time = "N/A"
                    if "full" in working:
                        intern_full_time = "full-time"
                    elif "intern" in working:
                        intern_full_time = "intern"
                    print "intern_full_time:", intern_full_time
                    r = re.findall(r"\((.*?)\)", working)
                    years_of_working = "N/A"
                    if r:
                        years_of_working = r[0]
                    # print "years_of_working:", years_of_working


                    recommends = outlook = aprroves = "N/A"

                    recommends_soup = comment.find("div", "flex-grid recommends")
                    if recommends_soup:
                        divs = recommends_soup.findAll("div", "tightLt col span-1-3")
                        for div in divs:
                            text = div.getText()
                            if "Recommend" in text:
                                recommends = text
                            if "Outlook" in text:
                                outlook = text
                            if "CEO" in text:
                                aprroves = text
                    # print recommends, outlook, aprroves


                    span_soup = comment.find("span", "gdStars gdRatings sm margRt")
                    span_soup = span_soup.find("span", "value-title")
                    overall_rating = span_soup.get("title")
                    # print "overall_rating:", overall_rating

                    culture_and_values = work_life = senior_management = comp_and_benefits = career_opportunities = "N/A"
                    ul_soup = comment.find("ul", "undecorated")
                    if ul_soup:
                        lis = ul_soup.findAll("li")
                        for li in lis:
                            t = li.getText()
                            value = li.find("span", "notranslate notranslate_title gdBars gdRatings med").get("title")
                            if "Comp" in t:
                                culture_and_values = value
                            elif "Work" in t:
                                work_life = value
                            elif "Senior" in t:
                                senior_management = value
                            elif "Culture" in t:
                                comp_and_benefits = value
                            elif "Career" in t:
                                career_opportunities = value
                    # print culture_and_values, work_life, senior_management, comp_and_benefits, career_opportunities


                    h2_soup = comment.find("h2", "h2 summary strong tightTop")
                    comment_title = h2_soup.find("span", "summary").getText()
                    # print "comment_title:", comment_title

                    div_soup = comment.find("div", "tbl fill prosConsAdvice")
                    rows = div_soup.findAll("div", "row")
                    pros = cons = advice = "N/A"
                    for row in rows:
                        # t = row.find("div", "cell padRt padBot strong top p").getText()
                        # c = row.find("p").getText()
                        t = row.find("p", "tightVert").getText()
                        c = row.find("p", "noMargVert").getText()
                        if "Pros" in t:
                            pros = c
                        elif "Cons" in t:
                            cons = c
                        elif "Advice" in t:
                            advice = c

                    # print "pros:", pros
                    # print "cons:", cons
                    # print "advice:", advice


                    print 

                    sheet2["A%d"%j].value = name
                    sheet2["B%d"%j].value = stock_code
                    sheet2["C%d"%j].value = n
                    sheet2["D%d"%j].value = comment_time
                    sheet2["E%d"%j].value = reviews
                    sheet2["F%d"%j].value = overallRating
                    sheet2["G%d"%j].value = cultureAndValues
                    sheet2["H%d"%j].value = workLife
                    sheet2["I%d"%j].value = seniorManagement
                    sheet2["J%d"%j].value = compAndBenefits
                    sheet2["K%d"%j].value = careerOpportunities
                    sheet2["L%d"%j].value = former_current
                    sheet2["M%d"%j].value = job_title
                    sheet2["N%d"%j].value = intern_full_time
                    sheet2["O%d"%j].value = years_of_working
                    sheet2["P%d"%j].value = overall_rating
                    sheet2["Q%d"%j].value = culture_and_values
                    sheet2["R%d"%j].value = work_life
                    sheet2["S%d"%j].value = senior_management
                    sheet2["T%d"%j].value = comp_and_benefits
                    sheet2["U%d"%j].value = career_opportunities
                    sheet2["V%d"%j].value = recommends
                    sheet2["W%d"%j].value = outlook
                    sheet2["X%d"%j].value = aprroves
                    sheet2["Y%d"%j].value = comment_title
                    sheet2["Z%d"%j].value = pros
                    sheet2["AA%d"%j].value = cons
                    sheet2["AB%d"%j].value = advice

                    j += 1

                    n += 1


                wb2.save("output.xlsx")

                open('finish_page.txt', 'a').write(p_url + '\n')

                if "<li class='next'>" not in p:
                    break
                if "<li class='next'> <span class='disabled'><i>" in p:
                    break

            except:
                traceback.print_exc()
                print 'error_page', p_url
                open('error_page.txt', 'a').write(p_url + '\n')

        if not flag:
            open("zero.txt", "a").write(name + "\n")

        open('finish_cmp.txt', 'a').write(url + '\n')
                    
except:
    traceback.print_exc()

print "Finish"
raw_input("")
